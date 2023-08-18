import openpyxl

# Load the Excel file
file_path = 'data.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Add a new column header for "Average"
sheet['D1'] = 'Average'

# Calculate average scores and update the "Average" column
for row in range(2, sheet.max_row + 1):
    math_score = sheet.cell(row=row, column=2).value
    science_score = sheet.cell(row=row, column=3).value
    average = (math_score + science_score) / 2
    sheet.cell(row=row, column=4, value=average)

# Save the updated Excel file
workbook.save(file_path)
workbook.close()

print("Average scores calculated and written to the Excel file.")
